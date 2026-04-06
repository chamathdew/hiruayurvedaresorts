import openpyxl
from openpyxl.styles import PatternFill
import json
from datetime import datetime

wb = openpyxl.load_workbook(r'hiruom.xlsx')
ws = wb.active

print("Sheets:", wb.sheetnames)
print("Max row:", ws.max_row, "Max col:", ws.max_column)
print()

months = []
current_month = None
current_section = None

all_rows = list(ws.iter_rows())

i = 0
while i < len(all_rows):
    row = all_rows[i]
    cell_a = row[0]
    val_a = cell_a.value

    # Check if this is a month header row (datetime in col A)
    if isinstance(val_a, datetime):
        if current_month:
            months.append(current_month)
        current_month = {
            'month': val_a.strftime('%Y-%m'),
            'sections': []
        }
        current_section = None
        i += 1
        continue

    # Check if section header
    if isinstance(val_a, str) and val_a.strip() in ['Ground Floor', 'First Floor', 'Reserve', 'Behind ', 'Roof Top', 'Corner 9', 'Behind']:
        current_section = val_a.strip()
        if current_month is not None:
            current_month['sections'].append({'name': current_section, 'rooms': []})
        i += 1
        continue

    # Check if room row (integer room number 1-20)
    if isinstance(val_a, (int, float)) and val_a == int(val_a) and 1 <= int(val_a) <= 20:
        room_no = int(val_a)
        bookings = []
        for col_idx, cell in enumerate(row[1:32], 2):  # cols B to AF (day 1-31)
            if cell.value and str(cell.value).strip():
                fgColor = cell.fill.fgColor.rgb if cell.fill and cell.fill.fgColor else '00000000'
                bookings.append({
                    'day': col_idx - 1,
                    'text': str(cell.value).strip(),
                    'color': fgColor
                })

        if current_month and current_month['sections']:
            current_month['sections'][-1]['rooms'].append({
                'room': room_no,
                'bookings': bookings
            })
        elif current_month:
            # No section yet, create default
            current_month['sections'].append({'name': 'General', 'rooms': []})
            current_month['sections'][-1]['rooms'].append({
                'room': room_no,
                'bookings': bookings
            })
        i += 1
        continue

    i += 1

if current_month:
    months.append(current_month)

# Print summary for first 3 months
for m in months[:4]:
    print("Month:", m['month'])
    for sec in m['sections']:
        print("  Section:", sec['name'])
        for rm in sec['rooms']:
            print("    Room", rm['room'], ":", len(rm['bookings']), "bookings")
            for b in rm['bookings'][:3]:
                print("      Day", b['day'], ":", b['text'][:60], "[color=" + b['color'] + "]")
    print()

# Also dump colors used
all_colors = set()
for m in months:
    for sec in m['sections']:
        for rm in sec['rooms']:
            for b in rm['bookings']:
                all_colors.add(b['color'])
print("All colors found:", all_colors)

# Save to JSON for reference
with open("tools/xlsx_data.json", "w", encoding="utf-8") as f:
    json.dump(months, f, ensure_ascii=False, indent=2)
print("Saved to tools/xlsx_data.json")
